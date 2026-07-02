import csv
import io

from flask import Blueprint, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from db.database import get_db

bp = Blueprint("export", __name__, url_prefix="/api/export")

COLUMNS = [
    ("name", "Agency"), ("city", "City"), ("deals_count", "Deals"),
    ("exclusives_count", "Exclusives"), ("phone_used", "Phone used"),
    ("phone_source", "Phone source"), ("website_url", "Website"),
    ("has_website", "Has website"), ("scraped_at", "Scraped date"),
    ("status", "Status"), ("sent_at", "Sent at"), ("replied_at", "Replied at"),
    ("notes", "Notes"), ("direct_mobile", "Direct mobile"),
]


def _fetch_rows():
    conn = get_db()
    try:
        return conn.execute("SELECT * FROM agencies ORDER BY scraped_at DESC").fetchall()
    finally:
        conn.close()


@bp.route("/xlsx", methods=["GET"])
def export_xlsx():
    rows = _fetch_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Madlan CRM"

    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in rows:
        ws.append([row[key] for key, _ in COLUMNS])
        if not row["has_website"]:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = yellow

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="madlan_crm.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/csv", methods=["GET"])
def export_csv():
    rows = _fetch_rows()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in COLUMNS])
    for row in rows:
        writer.writerow([row[key] for key, _ in COLUMNS])

    data = buf.getvalue().encode("utf-8-sig")  # BOM so Hebrew renders correctly in Excel
    return Response(
        data,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=madlan_crm.csv"},
    )
